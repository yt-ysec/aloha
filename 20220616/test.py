import os
import numpy as np
from PIL import Image
import openpyxl

wb = openpyxl.load_workbook('test.xlsx')
ws = wb['Sheet']

path = './mnist_mini_1000'
files = os.listdir(path)

imgDone = 0
kotae = 0

for file in files:
	if 'num4' in file or 'num9' in file:
		if 'num4' in file:
			kotae = 4
		else:
			kotae = 9

		imgPath = os.path.join(path, file)
		img = Image.open(imgPath)
		imgArray = np.array(img)

		for i in range(9):
			for j in range(9):
				ws.cell(row=1+i, column=1+j+9*imgDone).value = imgArray[i][j]
		ws.cell(row=10, column=1+9*imgDone).value = kotae
		imgDone = imgDone + 1

wb.save('test.xlsx')
